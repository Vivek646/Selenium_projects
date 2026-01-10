import openpyxl
path = "E:\worksheet.xlsx"

# writing the data into the sheet
workbook = openpyxl.load_workbook(path)
sheet = workbook['Sheet1']
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value = input('enter the data: ')
workbook.save(path)


# reading the data from the sheet
workbook = openpyxl.load_workbook(path)
sheet = workbook['Sheet1']
rows = sheet.max_row
cols = sheet.max_column
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value, end=' ')
    print()